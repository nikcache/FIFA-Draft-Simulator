#player Class

import openpyxl
from openpyxl import *
import re

class pDatabase:

    def __init__(self):

         #Opening workbook
        self.wb = load_workbook("data.xlsx") #, read_only = True)
        
        #Making sheet active
        self.ws = self.wb["data"]

    def getID(self, num):
        return self.ws.cell(row = (num + 1), column = 1).value

    def getName(self, num):
        return self.ws.cell(row = (num + 1), column = 3).value

    def getAge(self, num):
        return self.ws.cell(row = (num + 1), column = 4).value

    def getPhoto(self, num):
        return self.ws.cell(row = (num + 1), column = 5).value

    def getNation(self, num):
        return self.ws.cell(row = (num + 1), column = 6).value

    def getFlag(self, num):
        return self.ws.cell(row = (num + 1), column = 7).value

    def getOverall(self, num):
        return self.ws.cell(row = (num + 1), column = 8).value

    def getPotential(self, num):
        return self.ws.cell(row = (num + 1), column = 9).value

    def getClub(self, num):
        return self.ws.cell(row = (num + 1), column = 10).value

    def getClubLogo(self, num):
        return self.ws.cell(row = (num + 1), column = 11).value

    def getValue(self, num):
        return self.ws.cell(row = (num + 1), column = 12).value

    def getValueInt(self, num):
        try:
            val = self.ws.cell(row = (num + 1), column = 12).value
            
            valInt = float("".join(filter(lambda d: str.isdigit(d) or d == '.', val)))
            
            if val[-1] == "M":
                valInt = valInt * 1000000
            elif val[-1] == "K":
                valInt = valInt * 1000
            
            return int(valInt)

        except:
            return 0 
                   
    def getWage(self, num):
        return self.ws.cell(row = (num + 1), column = 13).value

    def getFoot(self, num):
        return self.ws.cell(row = (num + 1), column = 15).value

    def getIntRating(self, num):
        return self.ws.cell(row = (num + 1), column = 16).value

    def getWFoot(self, num):
        return self.ws.cell(row = (num + 1), column = 17).value

    def getSkill(self, num):
        return self.ws.cell(row = (num + 1), column = 18).value

    def getPosition(self, num):
        return self.ws.cell(row = (num + 1), column = 22).value

    def getNumber(self, num):
        return self.ws.cell(row = (num + 1), column = 23).value

    def getHeight(self, num):
        return self.ws.cell(row = (num + 1), column = 27).value

    def getWeight(self, num):
        return self.ws.cell(row = (num + 1), column = 27).value

    def getCrossing(self, num):
        return self.ws.cell(row = (num + 1), column = 54).value

    def getFinishing(self, num):
        return self.ws.cell(row = (num + 1), column = 55).value

    def getHeading(self, num):
        return self.ws.cell(row = (num + 1), column = 56).value

    def getPass(self, num):
        return self.ws.cell(row = (num + 1), column = 57).value

    def getVolley(self, num):
        return self.ws.cell(row = (num + 1), column = 58).value

    def getDribbling(self, num):
        return self.ws.cell(row = (num + 1), column = 59).value

    def getCurve(self, num):
        return self.ws.cell(row = (num + 1), column = 60).value

    def getFreeKick(self, num):
        return self.ws.cell(row = (num + 1), column = 61).value

    def getLongPass(self, num):
        return self.ws.cell(row = (num + 1), column = 62).value

    def getBallControl(self, num):
        return self.ws.cell(row = (num + 1), column = 63).value

    def getAcceleration(self, num):
        return self.ws.cell(row = (num + 1), column = 64).value

    def getSprintSpeed(self, num):
        return self.ws.cell(row = (num + 1), column = 65).value

    def getAgility(self, num):
        return self.ws.cell(row = (num + 1), column = 66).value

    def getReactions(self, num):
        return self.ws.cell(row = (num + 1), column = 68).value

    def getBalance(self, num):
        return self.ws.cell(row = (num + 1), column = 69).value

    def getShotPower(self, num):
        return self.ws.cell(row = (num + 1), column = 70).value

    def getJumping(self, num):
        return self.ws.cell(row = (num + 1), column = 71).value

    def getStamina(self, num):
        return self.ws.cell(row = (num + 1), column = 72).value

    def getStrength(self, num):
        return self.ws.cell(row = (num + 1), column = 73).value

    def getLongShots(self, num):
        return self.ws.cell(row = (num + 1), column = 74).value

    def getAggression(self, num):
        return self.ws.cell(row = (num + 1), column = 75).value

    def getInterception(self, num):
        return self.ws.cell(row = (num + 1), column = 76).value

    def getPositioning(self, num):
        return self.ws.cell(row = (num + 1), column = 77).value

    def getVision(self, num):
        return self.ws.cell(row = (num + 1), column = 78).value

    def getPenalties(self, num):
        return self.ws.cell(row = (num + 1), column = 79).value

    def getComposure(self, num):
        return self.ws.cell(row = (num + 1), column = 80).value

    def getMarking(self, num):
        return self.ws.cell(row = (num + 1), column = 81).value

    def getStandingTackle(self, num):
        return self.ws.cell(row = (num + 1), column = 82).value

    def getSlidingTackle(self, num):
        return self.ws.cell(row = (num + 1), column = 83).value

    def getGKDiving(self, num):
        return self.ws.cell(row = (num + 1), column = 84).value

    def getGKHandling(self, num):
        return self.ws.cell(row = (num + 1), column = 85).value

    def getGKKicking(self, num):
        return self.ws.cell(row = (num + 1), column = 86).value

    def getGKPositioning(self, num):
        return self.ws.cell(row = (num + 1), column = 87).value

    def getGKReflexes(self, num):
        return self.ws.cell(row = (num + 1), column = 88).value

    def getRelease(self, num):
        return self.ws.cell(row = (num + 1), column = 89).value
    
    def close(self):
        self.wb.close()

def main():
    player = pDatabase()
    for i in range(0, 100):
        print(player.getNation(i))
    player.close()

#main()
