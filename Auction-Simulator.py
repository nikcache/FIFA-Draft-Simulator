# Fifa Draft Simulator
# Nikesh Ghimire
# COM313 - Algorithmic Game Theory
# Mid-term

#Code is very inefficient - only made for purpose of the Class


#Importing Modules
from graphics import *
from txt import *
from buttonclass import *
import openpyxl
from openpyxl import *
from player import *
import player
from imageDL import *
import lxml
import random as rand
from random import *
from myTeam import *

#Loading Files Window
loadWin = GraphWin("Loading", 200, 200)

#App Logo
bannerImg = Image(Point(100, 75), "pics/Banner-03.png")
bannerImg.draw(loadWin)

#Text
txt("Loading...", 100, 175, loadWin, 20, "black", "bold")

#Opening Database
player = pDatabase()

#Opening user database
try:
    wb = load_workbook("userInfo.xlsx")
except FileNotFoundError:
    wb = Workbook()
    wb.create_sheet("User Info")
    wb.create_sheet("Sold Players")

    ws = wb["User Info"]
    sList = wb["Sold Players"]

    ws["A1"] = "Username"
    ws["B1"] = "PIN"

    sList["A1"] = "Player ID"
    sList["B1"] = "Name of Player"
    sList["C1"] = "Sold to:"
    sList["D1"] = "Amount"

loadWin.close()

#Defining starting function (initiates program)
def Intro():

    #Creating window 
    introWin = GraphWin("Intro Window", 300, 300)
    introWin.setBackground("#EDEEC9")

    ##Creating UI Elements in Intro Window

    #Banner
    bannerImg = Image(Point(150, 60), "pics/Banner-03.png")
    bannerImg.draw(introWin)

    #Text
    txt("Welcome to\nFIFA draft Simulator", 150, 135, introWin, 16, "#343E3D", "bold")
    txt("Created by:\nNikesh Ghimire", 150, 185, introWin, 12, "#343E3D", "normal")
    startButton = Button(introWin, Point(150, 235), 100, 25, "#DE6E4B", "Click to Start")
    quitButton = Button(introWin, Point(150, 265), 40, 25, "#DE6E4B", "Quit")

    #Button Clicks
    pt = introWin.getMouse()

    while not quitButton.isClicked(pt):

        if (startButton.isClicked(pt)):
            introWin.close()
            Login()

        try:        
            pt = introWin.getMouse()
        except GraphicsError:
            return

    introWin.close()

#Login window
def Login():

    databaseCheck()

    #Creating window
    logWin = GraphWin("Sign Up/Log In", 500, 300)
    logWin.setBackground("#EDEEC9")

    #Creating UI Elements in Login Window

    #BG
    rect1 = Rectangle(Point(0, 0), Point(500, 50))
    rect1.setFill("#031D44")
    rect1.draw(logWin)
    
    #Text
    txt("Sign Up/Log In", 90, 25, logWin, 16, "white", "bold")
    txt("Username:", 90, 100, logWin, 12, "#343E3D", "normal")
    txt("PIN:", 113, 150, logWin, 12, "#343E3D", "normal")
    txt("Note: If this is your first time (Sign Up) then remember your information!", 250, 275, logWin, 12, "#343E3D", "normal")

    #Dyanmic Text
    status = Text(Point(425, 25), "")
    status.setFill("white")
    status.draw(logWin)
    
    #Entry Boxes

    #PIN
    pinInput = Entry(Point(300, 150), 35)
    pinInput.draw(logWin)

    #Username
    userInput = Entry(Point(300, 100), 35)
    userInput.draw(logWin)

    #Buttons
    logButton = Button(logWin, Point(250, 200), 90, 25, "#DE6E4B", "Login")
    quitButton = Button(logWin, Point(250, 240), 40, 25, "#DE6E4B", "Quit")

    #Button Clicks
    pt = logWin.getMouse()

    while not (quitButton.isClicked(pt)):

        if (logButton.isClicked(pt)):

            if (check(userInput.getText(), pinInput.getText())):
                logWin.close()
                Menu(userInput.getText())
            else:
                status.setText("Incorrect Password\nTry Again")
    
        try:
            pt = logWin.getMouse()
        except GraphicsError:
            return
            
    logWin.close()

def Menu(name):

    userSheet = wb[name]
    #Creating Menu Window
    mwin = GraphWin("Menu", 300, 400)

    #Creating UI Elements in Menu Window

    #UI Design
    rect1 = Rectangle(Point(0, 0), Point(300, 50))
    rect1.setFill("#031D44")
    rect1.draw(mwin)
    
    #Text
    txt("Welcome "+name+"!", 150, 25, mwin, 16, "white", "bold")
    
    txt("What would you like to do today?", 150, 75, mwin, 12, "black", "normal")

    #Dynamic Text
    userBalance = Text(Point(150, 110), "Your Balance:\n€ " + str(round(userSheet["B2"].value)))
    userBalance.draw(mwin)
    
    #Buttons
    aucButton = Button(mwin, Point(150, 150), 125, 25, "#DE6E4B", "Draft Players")
    teamButton = Button(mwin, Point(150, 200), 125, 25, "#DE6E4B", "My Team")
    allButton = Button(mwin, Point(150, 250), 125, 25, "#DE6E4B", "See All Players")
    setButton = Button(mwin, Point(150, 300), 125, 25, "#DE6E4B", "Settings")
    quitButton = Button(mwin, Point(150, 350), 125, 25, "#DE6E4B", "Quit")
    outButton = Button(mwin, Point(230, 350), 25, 25, "#DE6E4B", "")

    #Editing buttons
    setButton.deactivate()

    #Icons
    logImage = Image(Point(230, 350), "pics/logout-02.png")
    logImage.draw(mwin)

    #Button Clicks

    pt = mwin.getMouse()

    while not (quitButton.isClicked(pt)):

        if aucButton.isClicked(pt):
            Auction(name)
            
        elif teamButton.isClicked(pt):
            MyTeam(name)
            
        elif allButton.isClicked(pt):
            PlayerList(name, 1)

        elif setButton.isClicked(pt):
            Settings()
            
        elif outButton.isClicked(pt):
            mwin.close()
            Login()

        userBalance.setText("Your Balance:\n€ " + str(userSheet["B2"].value))
        
        try:
            pt = mwin.getMouse()
        except GraphicsError:
            return

    
    mwin.close()

def Auction(name):

    #Creating a graphic Window
    awin = GraphWin("Auction Window", 600, 300)
    
    #Making player's sheet active in userInfo.xlsx
    userSheet = wb[name]
    histPlayers = wb["Sold Players"]

    #Drawing Objects
    rect1 = Rectangle(Point(0,0), Point(600, 50))
    rect1.setFill("#031D44")
    rect1.draw(awin)

    #Random player picker
    x = randrange(0, 18207)

    while x in histPlayers["A"]:
        x = randrange(0, 18207)

    #Text
    txt("Auction Window", 100, 25, awin, 16, "white", "bold")
    txt("Auction Details:", 100, 75, awin, 14, "black", "bold")

    #Dynamic Text

    #Name of Player
    txt("Name:", 170, 100, awin, 12, "black", "bold")
    playerName = Text(Point(300, 100), player.getName(x))
    playerName.draw(awin)

    #ID of Player
    txt("ID:", 170, 125, awin, 12, "black", "bold")
    playerID = Text(Point(300, 125), player.getID(x) + 1)
    playerID.draw(awin)

    #Overall Rating of Player
    txt("Rating:", 170, 150, awin, 12, "black", "bold")
    playerOverall = Text(Point(300, 150), player.getOverall(x))
    playerOverall.draw(awin)

    #Market Value of Player
    txt("Market Value:", 170, 175, awin, 12, "black", "bold")
    playerValue = Text(Point(300, 175), player.getValue(x))
    playerValue.draw(awin)

    #Balance of user
    userBalance = Text(Point(150, 275), "Your Balance: € " + str(userSheet["B2"].value))
    userBalance.draw(awin)

    #Status text
    statusTxt = Text(Point(480, 225), "")
    statusTxt.draw(awin)
    
    #Buttons
    backButton = Button(awin, Point(540, 25), 100, 25, "#DE6E4B", "Go Back")
    bidButton = Button(awin, Point(540, 275), 100, 25, "#DE6E4B", "Place Bid")
    nextButton = Button(awin, Point(415, 25), 100, 25, "#DE6E4B", "Next Player")
    viewButton = Button(awin, Point(290, 25), 100, 25, "#DE6E4B", "View Player")

    #Entry Box
    valueIn = Entry(Point(425, 275), 10)
    valueIn.draw(awin)

    #Button Clicks
    pt = awin.getMouse()
    while not (backButton.isClicked(pt)):

        if nextButton.isClicked(pt):

            statusTxt.setText("")
            bidButton.activate()
            x = randrange(0, 18207)

            while x in histPlayers["A"]:
                x = randrange(0, 18207)

            playerName.setText(player.getName(x))
            playerID.setText(player.getID(x) + 1)
            playerOverall.setText(player.getOverall(x))
            playerValue.setText(player.getValue(x))

        if viewButton.isClicked(pt):
            PlayerList(name, x)

        if bidButton.isClicked(pt):
            try:
                if int(valueIn.getText()) < userSheet["B2"].value: 
                    bidButton.deactivate()
                    playerAmount = playerWin(int(valueIn.getText()), x)

                    if playerAmount:
                        
                        emptyRowHist = histPlayers.max_row + 1
                        
                        histPlayers.cell(row = emptyRowHist, column = 1).value = player.getID(x) + 1
                        histPlayers.cell(row = emptyRowHist, column = 2).value = player.getName(x)
                        histPlayers.cell(row = emptyRowHist, column = 3).value = name
                        histPlayers.cell(row = emptyRowHist, column = 4).value = playerAmount

                        emptyRowUser = userSheet.max_row + 1

                        userSheet.cell(row = emptyRowUser, column = 1).value = player.getID(x) + 1
                        userSheet.cell(row = emptyRowUser, column = 2).value = player.getName(x)
                        userSheet.cell(row = emptyRowUser, column = 3).value = playerAmount

                        userSheet["B2"].value = userSheet["B2"].value - playerAmount
                        
                        saveBook()

                        statusTxt.setText("You won the player!\nYou paid € " + str(playerAmount))
                        userBalance.setText("Your Balance: € " + str(userSheet["B2"].value))
                        
                    else:
                        
                        statusTxt.setText("You didn't win the player!")
                else:
                    
                    statusTxt.setText("You don't have enough Money!")
                    
            except:
                statusTxt.setText("Error occured! Cannot value Player")
                
        try:
            pt = awin.getMouse()
        except GraphicsError:
            return

    awin.close()

#Function to check if player won and for what value
def playerWin(bidVal, num):

    bidList = []

    #Difficulty Modifier
    dMod = 1.5

    #Max number of bidders
    numBidders = 50

    #Getting value of player
    playerValue = player.getValueInt(num)

    #Generating bidders
    for bids in range(randrange(1, numBidders)):
        inflationValue = rand.uniform(0.8, 1.2)
        bidList.append(randrange(round(0.5 * playerValue), round(inflationValue * playerValue)))

    bidList.sort(reverse = True)
    
    #Printing number of bidders in shell
    print()
    print("Number of bidders: " + str(len(bidList)))
    print()
    print(bidList)
    print()
    
    #Setting initial value of b2
    b2 = 0

    #Loop to check through list of bids
    for bids in bidList:
        if bidVal <= bids:
            return False
        if bids > b2:
            b2 = bids
        if bidVal >= dMod * b2:
            b2 = bidVal
    return b2

#Function to see GUI interface of player's team
def MyTeam(name):

    #Instance Variable
    noPlayers = False

    #Creating View Player Window
    twin = GraphWin("My Team", 800, 475)

    #Drawing Objects
    rect1 = Rectangle(Point(0,0), Point(800, 50))
    rect1.setFill("#031D44")
    rect1.draw(twin)

    userSheet = wb[name]
    maxRowCount = userSheet.max_row

    myTeam = []
    
    for i in range(4, maxRowCount + 1):
        try:
            val = int(userSheet["A"+str(i)].value)
            myTeam.append(val)
        except:
            pass

    playerRosterNum = 0
    try:
        num = myTeam[playerRosterNum]
    except:
        num = 1
        noPlayers = True

                   
    #Buttons
    backButton = Button(twin, Point(750, 25), 75, 25, "#DE6E4B", "Go Back")
    prevButton = Button(twin, Point(250, 25), 25, 25, "#DE6E4B", "") 
    nextButton = Button(twin, Point(280, 25), 25, 25, "#DE6E4B", "")
    sellButton = Button(twin, Point(500, 25), 75, 25, "#DE6E4B", "Quick Sell")

    #Text
    txt("Player Database", 100, 25, twin, 16, "white", "bold")

    #Text Color picker
    colorList = ["white", "black"]
    colorInd = 0

    #Images
    
    #Borders
    specialImg = Image(Point(650, 250), "pics/special.png")
    goldImg = Image(Point(650, 250), "pics/gold.png")
    silverImg = Image(Point(650, 250), "pics/silver.png")
    bronzeImg = Image(Point(650, 250), "pics/bronze.png")

    #List of borders
    borderList = [specialImg, goldImg, silverImg, bronzeImg]

    #Draw border
    drawNum = 0

    if int(player.getOverall(num)) >= 90:
        drawNum = 0
        colorInd = 0
    elif 80 <= int(player.getOverall(num)) < 90:
        drawNum = 1
        colorInd = 1
    elif 70 <= int(player.getOverall(num)) < 80:
        drawNum = 2
        colorInd = 1
    elif int(player.getOverall(num)) < 70:
        drawNum = 3
        colorInd = 1

    borderList[drawNum].draw(twin)

    #Face
    if (faceDL(player.getPhoto(num), player.getID(num))) != -1:
        faceImg = Image(Point(675, 177), "pics/players/" + str(player.getID(num)) + ".png")
        faceImg.draw(twin)
    else:
        faceImg = Image(Point(675, 177), "pics/players/blank.png")
        faceImg.draw(twin)

    #Flag
    if (flagDL(player.getFlag(num), player.getNation(num))) != -1:
        flagImg = Image(Point(580, 190), "pics/flags/" + player.getNation(num) + ".png")
        flagImg.draw(twin)
    else:
        flagImg = Image(Point(580, 190), "pics/flags/blank.png")
        flagImg.draw(twin)

    #Club Logo
    if (clubDL(player.getClubLogo(num), player.getClub(num))) != -1:
        clubImg = Image(Point(580, 230), "pics/clubs/" + player.getClub(num) + ".png")
        clubImg.draw(twin)
    else:
        clubImg = Image(Point(580, 230), "pics/clubs/blank.png")
        clubImg.draw(twin)

    #UI Images
    backImg = Image(Point(250, 25), "pics/back.png")
    backImg.draw(twin)

    nextImg = Image(Point(280, 25), "pics/next.png")
    nextImg.draw(twin)
    
    #Dynamic Text

    #ID
    txt("ID:", 580, 25, twin, 16, "white", "bold")
    playerID = Text(Point(650, 25), int(player.getID(num))+1)
    playerID.setFill("white")
    playerID.setSize(16)
    playerID.setStyle("bold")
    playerID.draw(twin)

    #Name
    txt("Name:", 175, 75, twin, 12, "black", "bold")
    playerName = Text(Point(300, 75), player.getName(num))
    playerName.draw(twin)

    #Age
    txt("Age:", 175, 100, twin, 12, "black", "bold")
    playerAge = Text(Point(300, 100), player.getAge(num))
    playerAge.draw(twin)

    #Nation
    txt("Nation:", 175, 125, twin, 12, "black", "bold")
    playerNation = Text(Point(300, 125), player.getNation(num))
    playerNation.draw(twin)    

    #Overall
    playerOverall = Text(Point(580, 135), player.getOverall(num))
    playerOverall.setSize(30)
    playerOverall.setFace("helvetica")
    playerOverall.setFill(colorList[colorInd])
    playerOverall.setStyle("bold")
    playerOverall.draw(twin)

    #Name on card
    playerNameC = Text(Point(650, 267), player.getName(num))
    playerNameC.setSize(16)
    playerNameC.setFace("helvetica")
    playerNameC.setFill(colorList[colorInd])
    playerNameC.setStyle("bold")
    playerNameC.draw(twin)

    #Position
    playerPosition = Text(Point(580, 160), player.getPosition(num))
    playerPosition.setSize(16)
    playerPosition.setFace("helvetica")
    playerPosition.setFill(colorList[colorInd])
    playerPosition.setStyle("bold")
    playerPosition.draw(twin)

    if player.getPosition(num) != "GK":
        #Pace on Card
        player2 = Text(Point(595, 300), str(player.getAcceleration(num)) + " PAC")
        player2.setSize(14)
        player2.setFace("helvetica")
        player2.setFill(colorList[1])
        player2.setStyle("bold")
        player2.draw(twin)

        #Shot on Card
        player3 = Text(Point(595, 325), str(player.getFinishing(num)) + " SHO")
        player3.setSize(14)
        player3.setFace("helvetica")
        player3.setFill(colorList[1])
        player3.setStyle("bold")
        player3.draw(twin)

        #Pass on Card
        player4 = Text(Point(595, 350), str(player.getPass(num)) + " PAS")
        player4.setSize(14)
        player4.setFace("helvetica")
        player4.setFill(colorList[1])
        player4.setStyle("bold")
        player4.draw(twin)

        #Pace on Card
        player5 = Text(Point(705, 300), str(player.getDribbling(num)) + " DRI ")
        player5.setSize(14)
        player5.setFace("helvetica")
        player5.setFill(colorList[1])
        player5.setStyle("bold")
        player5.draw(twin)

        #Defending on Card
        player6 = Text(Point(705, 325), str(player.getInterception(num)) + " DEF")
        player6.setSize(14)
        player6.setFace("helvetica")
        player6.setFill(colorList[1])
        player6.setStyle("bold")
        player6.draw(twin)

        #Strength on Card
        player7 = Text(Point(705, 350), str(player.getStrength(num)) + " PHY")
        player7.setSize(14)
        player7.setFace("helvetica")
        player7.setFill(colorList[1])
        player7.setStyle("bold")
        player7.draw(twin)
    else:
        #GKDiving on Card
        player2 = Text(Point(595, 300), str(player.getGKDiving(num)) + " DIV ")
        player2.setSize(14)
        player2.setFace("helvetica")
        player2.setFill(colorList[1])
        player2.setStyle("bold")
        player2.draw(twin)

        #GKHandling on Card
        player3 = Text(Point(595, 325), str(player.getGKHandling(num)) + " HAN")
        player3.setSize(14)
        player3.setFace("helvetica")
        player3.setFill(colorList[1])
        player3.setStyle("bold")
        player3.draw(twin)

        #GKKicking on Card
        player4 = Text(Point(595, 350), str(player.getGKKicking(num)) + " KIC ")
        player4.setSize(14)
        player4.setFace("helvetica")
        player4.setFill(colorList[1])
        player4.setStyle("bold")
        player4.draw(twin)

        #GKReflexes on Card
        player5 = Text(Point(705, 300), str(player.getGKReflexes(num)) + " REF")
        player5.setSize(14)
        player5.setFace("helvetica")
        player5.setFill(colorList[1])
        player5.setStyle("bold")
        player5.draw(twin)

        #Agility on Card
        player6 = Text(Point(705, 325), str(player.getAgility(num)) + " SPE")
        player6.setSize(14)
        player6.setFace("helvetica")
        player6.setFill(colorList[1])
        player6.setStyle("bold")
        player6.draw(twin)

        #GKPositioning on Card
        player7 = Text(Point(705, 350), str(player.getGKPositioning(num)) + " POS")
        player7.setSize(14)
        player7.setFace("helvetica")
        player7.setFill(colorList[1])
        player7.setStyle("bold")
        player7.draw(twin)

    #Club
    txt("Club:", 175, 150, twin, 12, "black", "bold")
    playerClub = Text(Point(300, 150), player.getClub(num))
    playerClub.draw(twin)

    #Potential
    txt("Potential:", 175, 175, twin, 12, "black", "bold")
    playerPotential = Text(Point(300, 175), player.getPotential(num))
    playerPotential.draw(twin)

    #Value
    txt("Value:", 175, 200, twin, 12, "black", "bold")
    playerValue = Text(Point(300, 200), player.getValue(num))
    playerValue.draw(twin)

    #Wage
    txt("Wage:", 175, 225, twin, 12, "black", "bold")
    playerWage = Text(Point(300, 225), player.getWage(num))
    playerWage.draw(twin)

    #Strong Foot
    txt("Strong Foot:", 175, 250, twin, 12, "black", "bold")
    playerFoot = Text(Point(300, 250), player.getFoot(num))
    playerFoot.draw(twin)

    #Weak Foot
    txt("WeakFoot:", 175, 275, twin, 12, "black", "bold")
    playerWFoot = Text(Point(300, 275), str(player.getWFoot(num))+"/5")
    playerWFoot.draw(twin)

    #Weak Foot
    txt("Skill Rating:", 175, 300, twin, 12, "black", "bold")
    playerSkill = Text(Point(300, 300), str(player.getSkill(num))+"/5")
    playerSkill.draw(twin)

    #Internation Rating
    txt("Int Rating:", 175, 325, twin, 12, "black", "bold")
    playerIntRating = Text(Point(300, 325), str(player.getIntRating(num))+"/5")
    playerIntRating.draw(twin)

    #Jersey Number
    txt("Jersey Number:", 175, 350, twin, 12, "black", "bold")
    playerNumber = Text(Point(300, 350), player.getNumber(num))
    playerNumber.draw(twin)

    #Height
    txt("Height:", 175, 375, twin, 12, "black", "bold")
    playerHeight = Text(Point(300, 375), player.getHeight(num))
    playerHeight.draw(twin)

    #Weight
    txt("Weight:", 175, 400, twin, 12, "black", "bold")
    playerWeight = Text(Point(300, 400), player.getWeight(num))
    playerWeight.draw(twin)

    #Release Clause 
    txt("Release Clause:", 175, 425, twin, 12, "black", "bold")
    playerRelease = Text(Point(300, 425), player.getRelease(num))
    playerRelease.draw(twin)

    #Quick Sell Values
    multiplier = 0.95
    newVal = round(multiplier * player.getValueInt(num))
    sellValue = Text(Point(400, 25), "€ " + str(newVal))
    sellValue.setFill("white")
    sellValue.setStyle("bold")
    sellValue.draw(twin)
    
    #Checker
    if noPlayers:
            sellValue.setSize(10)
            sellValue.setText("No more players!\nClick to Close!")
            pt = twin.getMouse()
            try:
                twin.close()
                return
            except:
                return
                
    #Internal Variables
    try:
        if 1 <= num <= 18207:
            pID = True
        else:
            pID = False
    except:
        num = 1
        playerID.setText("Invalid ID")
                     
    #Button Clicks
    try:
        pt = twin.getMouse()
    except:
        return

    while not (backButton.isClicked(pt)):
        
        if (prevButton.isClicked(pt)):
            if playerRosterNum - 1 < 0:
                playerRosterNum = len(myTeam) - 1
            else:
                playerRosterNum = playerRosterNum - 1
            num = myTeam[playerRosterNum]
            
            
        elif (nextButton.isClicked(pt)):
            if playerRosterNum + 1 > len(myTeam) - 1:
                playerRosterNum = 0
            else:
                playerRosterNum = playerRosterNum + 1
            num = myTeam[playerRosterNum]
            

        elif (sellButton.isClicked(pt)):
            try:
                for cell in userSheet["A"]:
                    if cell.value == num:
                        cell.value = "sold"

                userSheet["B2"].value = userSheet["B2"].value + newVal        

                for item in myTeam:
                    if item == num:
                        myTeam.remove(item)

                if playerRosterNum + 1 > len(myTeam) - 1:
                    playerRosterNum = 0
                else:
                    playerRosterNum = playerRosterNum + 1
                    
                num = myTeam[playerRosterNum]
                
            except:
                num = 1
                sellValue.setSize(10)
                sellValue.setText("No more players!\nClick to close!")
                saveBook()
                twin.getMouse()
                try:
                    twin.close()
                    return
                except:
                    return
                
            

        try:
            borderList[drawNum].undraw()
        except:
            return

        
        #Update Quick sell values
        multiplier = 0.95
        newVal = round(multiplier * player.getValueInt(num))
        sellValue.setText("€ " + str(newVal))


        #Border and font color picker
        if int(player.getOverall(num)) >= 90:
            drawNum = 0
            colorInd = 0
        elif 80 <= int(player.getOverall(num)) < 90:
            drawNum = 1
            colorInd = 1
        elif 70 <= int(player.getOverall(num)) < 80:
            drawNum = 2
            colorInd = 1
        elif int(player.getOverall(num)) < 70:
            drawNum = 3
            colorInd = 1

        #Draw Border
        borderList[drawNum].draw(twin)

        #Draw face
        if (faceDL(player.getPhoto(num), player.getID(num))) != -1:
            faceImg.undraw()
            faceImg = Image(Point(675, 177), "pics/players/" + str(player.getID(num)) + ".png")
            faceImg.draw(twin)
        else:
            faceImg.undraw()
            faceImg = Image(Point(675, 177), "pics/players/blank.png")
            faceImg.draw(twin)
            
        #On Card
        
        #Overall
        playerOverall.undraw()
        playerOverall = Text(Point(580, 135), player.getOverall(num))
        playerOverall.setSize(30)
        playerOverall.setFace("helvetica")
        playerOverall.setFill(colorList[colorInd])
        playerOverall.setStyle("bold")
        playerOverall.draw(twin)

        
        #Position
        playerPosition.undraw()
        playerPosition = Text(Point(580, 160), player.getPosition(num))
        playerPosition.setSize(16)
        playerPosition.setFace("helvetica")
        playerPosition.setFill(colorList[colorInd])
        playerPosition.setStyle("bold")
        playerPosition.draw(twin)
        
        if player.getPosition(num) != "GK":
            #Pace on Card
            player2.undraw()
            player2 = Text(Point(595, 300), str(player.getAcceleration(num)) + " PAC")
            player2.setSize(14)
            player2.setFace("helvetica")
            player2.setFill(colorList[1])
            player2.setStyle("bold")
            player2.draw(twin)

            #Shot on Card
            player3.undraw()
            player3 = Text(Point(595, 325), str(player.getFinishing(num)) + " SHO")
            player3.setSize(14)
            player3.setFace("helvetica")
            player3.setFill(colorList[1])
            player3.setStyle("bold")
            player3.draw(twin)

            #Pass on Card
            player4.undraw()
            player4 = Text(Point(595, 350), str(player.getPass(num)) + " PAS")
            player4.setSize(14)
            player4.setFace("helvetica")
            player4.setFill(colorList[1])
            player4.setStyle("bold")
            player4.draw(twin)

            #Pace on Card
            player5.undraw()
            player5 = Text(Point(705, 300), str(player.getDribbling(num)) + " DRI ")
            player5.setSize(14)
            player5.setFace("helvetica")
            player5.setFill(colorList[1])
            player5.setStyle("bold")
            player5.draw(twin)

            #Defending on Card
            player6.undraw()
            player6 = Text(Point(705, 325), str(player.getInterception(num)) + " DEF")
            player6.setSize(14)
            player6.setFace("helvetica")
            player6.setFill(colorList[1])
            player6.setStyle("bold")
            player6.draw(twin)

            #Strength on Card
            player7.undraw()
            player7 = Text(Point(705, 350), str(player.getStrength(num)) + " PHY")
            player7.setSize(14)
            player7.setFace("helvetica")
            player7.setFill(colorList[1])
            player7.setStyle("bold")
            player7.draw(twin)
        else:
            #GKDiving on Card
            player2.undraw()
            player2 = Text(Point(595, 300), str(player.getGKDiving(num)) + " DIV ")
            player2.setSize(14)
            player2.setFace("helvetica")
            player2.setFill(colorList[1])
            player2.setStyle("bold")
            player2.draw(twin)

            #GKHandling on Card
            player3.undraw()
            player3 = Text(Point(595, 325), str(player.getGKHandling(num)) + " HAN")
            player3.setSize(14)
            player3.setFace("helvetica")
            player3.setFill(colorList[1])
            player3.setStyle("bold")
            player3.draw(twin)

            #GKKicking on Card
            player4.undraw()
            player4 = Text(Point(595, 350), str(player.getGKKicking(num)) + " KIC ")
            player4.setSize(14)
            player4.setFace("helvetica")
            player4.setFill(colorList[1])
            player4.setStyle("bold")
            player4.draw(twin)

            #GKReflexes on Card
            player5.undraw()
            player5 = Text(Point(705, 300), str(player.getGKReflexes(num)) + " REF")
            player5.setSize(14)
            player5.setFace("helvetica")
            player5.setFill(colorList[1])
            player5.setStyle("bold")
            player5.draw(twin)

            #Agility on Card
            player6.undraw()
            player6 = Text(Point(705, 325), str(player.getAgility(num)) + " SPE")
            player6.setSize(14)
            player6.setFace("helvetica")
            player6.setFill(colorList[1])
            player6.setStyle("bold")
            player6.draw(twin)

            #GKPositioning on Card
            player7.undraw()
            player7 = Text(Point(705, 350), str(player.getGKPositioning(num)) + " POS")
            player7.setSize(14)
            player7.setFace("helvetica")
            player7.setFill(colorList[1])
            player7.setStyle("bold")
            player7.draw(twin)

        #Flag
        if (flagDL(player.getFlag(num), player.getNation(num))) != -1:
            flagImg.undraw() 
            flagImg = Image(Point(580, 190), "pics/flags/" + player.getNation(num) + ".png")
            flagImg.draw(twin)
        else:
            flagImg.undraw()
            flagImg = Image(Point(580, 190), "pics/flags/blank.png")
            flagImg.draw(twin)

        #Club
        if (clubDL(player.getClubLogo(num), player.getClub(num))) != -1:
            clubImg.undraw()
            clubImg = Image(Point(580, 230), "pics/clubs/" + player.getClub(num) + ".png")
            clubImg.draw(twin)
        else:
            clubImg.undraw()
            clubImg = Image(Point(580, 230), "pics/clubs/blank.png")
            clubImg.draw(twin)

        #Name on card
        playerNameC.undraw()
        playerNameC = Text(Point(650, 267), player.getName(num))
        playerNameC.setSize(16)
        playerNameC.setFace("helvetica")
        playerNameC.setFill(colorList[colorInd])
        playerNameC.setStyle("bold")
        playerNameC.draw(twin)
        

        #Printed Details
        if pID:
            playerID.setText(int(player.getID(num))+1)
        if noPlayers:
            sellValue.setSize(10)
            sellValue.setText("No more players!\nClick to Close!")
            pt = twin.getMouse()
            try:
                twin.close()
            except:
                twin.close()
        playerName.setText(player.getName(num))
        playerAge.setText(player.getAge(num))
        playerNation.setText(player.getNation(num))
        playerClub.setText(player.getClub(num))
        playerPotential.setText(player.getPotential(num))
        playerValue.setText(player.getValue(num))
        playerWage.setText(player.getWage(num))
        playerFoot.setText(player.getFoot(num))
        playerWFoot.setText(str(player.getWFoot(num))+"/5")
        playerSkill.setText(str(player.getSkill(num))+"/5")
        playerIntRating.setText(str(player.getIntRating(num))+"/5")
        playerNumber.setText(player.getNumber(num))
        playerHeight.setText(player.getHeight(num))
        playerWeight.setText(player.getWeight(num))
        playerRelease.setText(player.getRelease(num))
        
        try:
            pt = twin.getMouse()
        except GraphicsError:
            pass

    saveBook()
    twin.close()   

def PlayerList(name, num):

    #Creating View Player Window
    pwin = GraphWin("Player Database", 800, 475)

    #Drawing Objects
    rect1 = Rectangle(Point(0,0), Point(800, 50))
    rect1.setFill("#031D44")
    rect1.draw(pwin)

    #Buttons
    backButton = Button(pwin, Point(750, 25), 75, 25, "#DE6E4B", "Go Back")
    prevButton = Button(pwin, Point(250, 25), 25, 25, "#DE6E4B", "") 
    nextButton = Button(pwin, Point(280, 25), 25, 25, "#DE6E4B", "")
    jumpButton = Button(pwin, Point(500, 25), 75, 25, "#DE6E4B", "Search ID")

    #Entry Boxes
    idSearch = Entry(Point(430, 25), 6)
    idSearch.draw(pwin)

    #Text
    txt("Player Database", 100, 25, pwin, 16, "white", "bold")

    #Text Color picker
    colorList = ["white", "black"]
    colorInd = 0

    #Images
    
    #Borders
    specialImg = Image(Point(650, 250), "pics/special.png")
    goldImg = Image(Point(650, 250), "pics/gold.png")
    silverImg = Image(Point(650, 250), "pics/silver.png")
    bronzeImg = Image(Point(650, 250), "pics/bronze.png")

    #List of borders
    borderList = [specialImg, goldImg, silverImg, bronzeImg]

    #Draw border
    drawNum = 0

    if int(player.getOverall(num)) >= 90:
        drawNum = 0
        colorInd = 0
    elif 80 <= int(player.getOverall(num)) < 90:
        drawNum = 1
        colorInd = 1
    elif 70 <= int(player.getOverall(num)) < 80:
        drawNum = 2
        colorInd = 1
    elif int(player.getOverall(num)) < 70:
        drawNum = 3
        colorInd = 1

    borderList[drawNum].draw(pwin)

    #Face
    if (faceDL(player.getPhoto(num), player.getID(num))) != -1:
        faceImg = Image(Point(675, 177), "pics/players/" + str(player.getID(num)) + ".png")
        faceImg.draw(pwin)
    else:
        faceImg = Image(Point(675, 177), "pics/players/blank.png")
        faceImg.draw(pwin)

    #Flag
    if (flagDL(player.getFlag(num), player.getNation(num))) != -1:
        flagImg = Image(Point(580, 190), "pics/flags/" + player.getNation(num) + ".png")
        flagImg.draw(pwin)
    else:
        flagImg = Image(Point(580, 190), "pics/flags/blank.png")
        flagImg.draw(pwin)

    #Club Logo
    if (clubDL(player.getClubLogo(num), player.getClub(num))) != -1:
        clubImg = Image(Point(580, 230), "pics/clubs/" + player.getClub(num) + ".png")
        clubImg.draw(pwin)
    else:
        clubImg = Image(Point(580, 230), "pics/clubs/blank.png")
        clubImg.draw(pwin)

    #UI Images
    backImg = Image(Point(250, 25), "pics/back.png")
    backImg.draw(pwin)

    nextImg = Image(Point(280, 25), "pics/next.png")
    nextImg.draw(pwin)
    
    #Dynamic Text

    #ID
    txt("ID:", 580, 25, pwin, 16, "white", "bold")
    playerID = Text(Point(650, 25), int(player.getID(num))+1)
    playerID.setFill("white")
    playerID.setSize(16)
    playerID.setStyle("bold")
    playerID.draw(pwin)

    #Name
    txt("Name:", 175, 75, pwin, 12, "black", "bold")
    playerName = Text(Point(300, 75), player.getName(num))
    playerName.draw(pwin)

    #Age
    txt("Age:", 175, 100, pwin, 12, "black", "bold")
    playerAge = Text(Point(300, 100), player.getAge(num))
    playerAge.draw(pwin)

    #Nation
    txt("Nation:", 175, 125, pwin, 12, "black", "bold")
    playerNation = Text(Point(300, 125), player.getNation(num))
    playerNation.draw(pwin)    

    #Overall
    playerOverall = Text(Point(580, 135), player.getOverall(num))
    playerOverall.setSize(30)
    playerOverall.setFace("helvetica")
    playerOverall.setFill(colorList[colorInd])
    playerOverall.setStyle("bold")
    playerOverall.draw(pwin)

    #Name on card
    playerNameC = Text(Point(650, 267), player.getName(num))
    playerNameC.setSize(16)
    playerNameC.setFace("helvetica")
    playerNameC.setFill(colorList[colorInd])
    playerNameC.setStyle("bold")
    playerNameC.draw(pwin)

    #Position
    playerPosition = Text(Point(580, 160), player.getPosition(num))
    playerPosition.setSize(16)
    playerPosition.setFace("helvetica")
    playerPosition.setFill(colorList[colorInd])
    playerPosition.setStyle("bold")
    playerPosition.draw(pwin)

    if player.getPosition(num) != "GK":
        #Pace on Card
        player2 = Text(Point(595, 300), str(player.getAcceleration(num)) + " PAC")
        player2.setSize(14)
        player2.setFace("helvetica")
        player2.setFill(colorList[1])
        player2.setStyle("bold")
        player2.draw(pwin)

        #Shot on Card
        player3 = Text(Point(595, 325), str(player.getFinishing(num)) + " SHO")
        player3.setSize(14)
        player3.setFace("helvetica")
        player3.setFill(colorList[1])
        player3.setStyle("bold")
        player3.draw(pwin)

        #Pass on Card
        player4 = Text(Point(595, 350), str(player.getPass(num)) + " PAS")
        player4.setSize(14)
        player4.setFace("helvetica")
        player4.setFill(colorList[1])
        player4.setStyle("bold")
        player4.draw(pwin)

        #Pace on Card
        player5 = Text(Point(705, 300), str(player.getDribbling(num)) + " DRI ")
        player5.setSize(14)
        player5.setFace("helvetica")
        player5.setFill(colorList[1])
        player5.setStyle("bold")
        player5.draw(pwin)

        #Defending on Card
        player6 = Text(Point(705, 325), str(player.getInterception(num)) + " DEF")
        player6.setSize(14)
        player6.setFace("helvetica")
        player6.setFill(colorList[1])
        player6.setStyle("bold")
        player6.draw(pwin)

        #Strength on Card
        player7 = Text(Point(705, 350), str(player.getStrength(num)) + " PHY")
        player7.setSize(14)
        player7.setFace("helvetica")
        player7.setFill(colorList[1])
        player7.setStyle("bold")
        player7.draw(pwin)
    else:
        #GKDiving on Card
        player2 = Text(Point(595, 300), str(player.getGKDiving(num)) + " DIV ")
        player2.setSize(14)
        player2.setFace("helvetica")
        player2.setFill(colorList[1])
        player2.setStyle("bold")
        player2.draw(pwin)

        #GKHandling on Card
        player3 = Text(Point(595, 325), str(player.getGKHandling(num)) + " HAN")
        player3.setSize(14)
        player3.setFace("helvetica")
        player3.setFill(colorList[1])
        player3.setStyle("bold")
        player3.draw(pwin)

        #GKKicking on Card
        player4 = Text(Point(595, 350), str(player.getGKKicking(num)) + " KIC ")
        player4.setSize(14)
        player4.setFace("helvetica")
        player4.setFill(colorList[1])
        player4.setStyle("bold")
        player4.draw(pwin)

        #GKReflexes on Card
        player5 = Text(Point(705, 300), str(player.getGKReflexes(num)) + " REF")
        player5.setSize(14)
        player5.setFace("helvetica")
        player5.setFill(colorList[1])
        player5.setStyle("bold")
        player5.draw(pwin)

        #Agility on Card
        player6 = Text(Point(705, 325), str(player.getAgility(num)) + " SPE")
        player6.setSize(14)
        player6.setFace("helvetica")
        player6.setFill(colorList[1])
        player6.setStyle("bold")
        player6.draw(pwin)

        #GKPositioning on Card
        player7 = Text(Point(705, 350), str(player.getGKPositioning(num)) + " POS")
        player7.setSize(14)
        player7.setFace("helvetica")
        player7.setFill(colorList[1])
        player7.setStyle("bold")
        player7.draw(pwin)

    #Club
    txt("Club:", 175, 150, pwin, 12, "black", "bold")
    playerClub = Text(Point(300, 150), player.getClub(num))
    playerClub.draw(pwin)

    #Potential
    txt("Potential:", 175, 175, pwin, 12, "black", "bold")
    playerPotential = Text(Point(300, 175), player.getPotential(num))
    playerPotential.draw(pwin)

    #Value
    txt("Value:", 175, 200, pwin, 12, "black", "bold")
    playerValue = Text(Point(300, 200), player.getValue(num))
    playerValue.draw(pwin)

    #Wage
    txt("Wage:", 175, 225, pwin, 12, "black", "bold")
    playerWage = Text(Point(300, 225), player.getWage(num))
    playerWage.draw(pwin)

    #Strong Foot
    txt("Strong Foot:", 175, 250, pwin, 12, "black", "bold")
    playerFoot = Text(Point(300, 250), player.getFoot(num))
    playerFoot.draw(pwin)

    #Weak Foot
    txt("WeakFoot:", 175, 275, pwin, 12, "black", "bold")
    playerWFoot = Text(Point(300, 275), str(player.getWFoot(num))+"/5")
    playerWFoot.draw(pwin)

    #Weak Foot
    txt("Skill Rating:", 175, 300, pwin, 12, "black", "bold")
    playerSkill = Text(Point(300, 300), str(player.getSkill(num))+"/5")
    playerSkill.draw(pwin)

    #Internation Rating
    txt("Int Rating:", 175, 325, pwin, 12, "black", "bold")
    playerIntRating = Text(Point(300, 325), str(player.getIntRating(num))+"/5")
    playerIntRating.draw(pwin)

    #Jersey Number
    txt("Jersey Number:", 175, 350, pwin, 12, "black", "bold")
    playerNumber = Text(Point(300, 350), player.getNumber(num))
    playerNumber.draw(pwin)

    #Height
    txt("Height:", 175, 375, pwin, 12, "black", "bold")
    playerHeight = Text(Point(300, 375), player.getHeight(num))
    playerHeight.draw(pwin)

    #Weight
    txt("Weight:", 175, 400, pwin, 12, "black", "bold")
    playerWeight = Text(Point(300, 400), player.getWeight(num))
    playerWeight.draw(pwin)

    #Release Clause 
    txt("Release Clause:", 175, 425, pwin, 12, "black", "bold")
    playerRelease = Text(Point(300, 425), player.getRelease(num))
    playerRelease.draw(pwin)

    #Internal Variables
    try:
        if 1 <= num <= 18207:
            pID = True
        else:
            pID = False
    except:
        num = 1
        playerID.setText("Invalid ID")
    #Button Clicks
    pt = pwin.getMouse()

    while not (backButton.isClicked(pt)):

        if (prevButton.isClicked(pt)):
            if num >= 2:
                num = num - 1
                
        
        elif (nextButton.isClicked(pt)):
            if num <= 18206:
                num = num + 1

        elif (jumpButton.isClicked(pt)):
            try:
                newID = eval(idSearch.getText())
                if 1 <= newID <= 18207:
                    num = newID
                else:
                    playerID.setText("Invalid ID")
                    pID = False
            except:
                playerID.setText("Invalid ID")
                pID = False

        borderList[drawNum].undraw()

        #Border and font color picker
        if int(player.getOverall(num)) >= 90:
            drawNum = 0
            colorInd = 0
        elif 80 <= int(player.getOverall(num)) < 90:
            drawNum = 1
            colorInd = 1
        elif 70 <= int(player.getOverall(num)) < 80:
            drawNum = 2
            colorInd = 1
        elif int(player.getOverall(num)) < 70:
            drawNum = 3
            colorInd = 1

        #Draw Border
        borderList[drawNum].draw(pwin)

        #Draw face
        if (faceDL(player.getPhoto(num), player.getID(num))) != -1:
            faceImg.undraw()
            faceImg = Image(Point(675, 177), "pics/players/" + str(player.getID(num)) + ".png")
            faceImg.draw(pwin)
        else:
            faceImg.undraw()
            faceImg = Image(Point(675, 177), "pics/players/blank.png")
            faceImg.draw(pwin)
            
        #On Card
        
        #Overall
        playerOverall.undraw()
        playerOverall = Text(Point(580, 135), player.getOverall(num))
        playerOverall.setSize(30)
        playerOverall.setFace("helvetica")
        playerOverall.setFill(colorList[colorInd])
        playerOverall.setStyle("bold")
        playerOverall.draw(pwin)

        
        #Position
        playerPosition.undraw()
        playerPosition = Text(Point(580, 160), player.getPosition(num))
        playerPosition.setSize(16)
        playerPosition.setFace("helvetica")
        playerPosition.setFill(colorList[colorInd])
        playerPosition.setStyle("bold")
        playerPosition.draw(pwin)
        
        if player.getPosition(num) != "GK":
            #Pace on Card
            player2.undraw()
            player2 = Text(Point(595, 300), str(player.getAcceleration(num)) + " PAC")
            player2.setSize(14)
            player2.setFace("helvetica")
            player2.setFill(colorList[1])
            player2.setStyle("bold")
            player2.draw(pwin)

            #Shot on Card
            player3.undraw()
            player3 = Text(Point(595, 325), str(player.getFinishing(num)) + " SHO")
            player3.setSize(14)
            player3.setFace("helvetica")
            player3.setFill(colorList[1])
            player3.setStyle("bold")
            player3.draw(pwin)

            #Pass on Card
            player4.undraw()
            player4 = Text(Point(595, 350), str(player.getPass(num)) + " PAS")
            player4.setSize(14)
            player4.setFace("helvetica")
            player4.setFill(colorList[1])
            player4.setStyle("bold")
            player4.draw(pwin)

            #Pace on Card
            player5.undraw()
            player5 = Text(Point(705, 300), str(player.getDribbling(num)) + " DRI ")
            player5.setSize(14)
            player5.setFace("helvetica")
            player5.setFill(colorList[1])
            player5.setStyle("bold")
            player5.draw(pwin)

            #Defending on Card
            player6.undraw()
            player6 = Text(Point(705, 325), str(player.getInterception(num)) + " DEF")
            player6.setSize(14)
            player6.setFace("helvetica")
            player6.setFill(colorList[1])
            player6.setStyle("bold")
            player6.draw(pwin)

            #Strength on Card
            player7.undraw()
            player7 = Text(Point(705, 350), str(player.getStrength(num)) + " PHY")
            player7.setSize(14)
            player7.setFace("helvetica")
            player7.setFill(colorList[1])
            player7.setStyle("bold")
            player7.draw(pwin)
        else:
            #GKDiving on Card
            player2.undraw()
            player2 = Text(Point(595, 300), str(player.getGKDiving(num)) + " DIV ")
            player2.setSize(14)
            player2.setFace("helvetica")
            player2.setFill(colorList[1])
            player2.setStyle("bold")
            player2.draw(pwin)

            #GKHandling on Card
            player3.undraw()
            player3 = Text(Point(595, 325), str(player.getGKHandling(num)) + " HAN")
            player3.setSize(14)
            player3.setFace("helvetica")
            player3.setFill(colorList[1])
            player3.setStyle("bold")
            player3.draw(pwin)

            #GKKicking on Card
            player4.undraw()
            player4 = Text(Point(595, 350), str(player.getGKKicking(num)) + " KIC ")
            player4.setSize(14)
            player4.setFace("helvetica")
            player4.setFill(colorList[1])
            player4.setStyle("bold")
            player4.draw(pwin)

            #GKReflexes on Card
            player5.undraw()
            player5 = Text(Point(705, 300), str(player.getGKReflexes(num)) + " REF")
            player5.setSize(14)
            player5.setFace("helvetica")
            player5.setFill(colorList[1])
            player5.setStyle("bold")
            player5.draw(pwin)

            #Agility on Card
            player6.undraw()
            player6 = Text(Point(705, 325), str(player.getAgility(num)) + " SPE")
            player6.setSize(14)
            player6.setFace("helvetica")
            player6.setFill(colorList[1])
            player6.setStyle("bold")
            player6.draw(pwin)

            #GKPositioning on Card
            player7.undraw()
            player7 = Text(Point(705, 350), str(player.getGKPositioning(num)) + " POS")
            player7.setSize(14)
            player7.setFace("helvetica")
            player7.setFill(colorList[1])
            player7.setStyle("bold")
            player7.draw(pwin)

        #Flag
        if (flagDL(player.getFlag(num), player.getNation(num))) != -1:
            flagImg.undraw() 
            flagImg = Image(Point(580, 190), "pics/flags/" + player.getNation(num) + ".png")
            flagImg.draw(pwin)
        else:
            flagImg.undraw()
            flagImg = Image(Point(580, 190), "pics/flags/blank.png")
            flagImg.draw(pwin)

        #Club
        if (clubDL(player.getClubLogo(num), player.getClub(num))) != -1:
            clubImg.undraw()
            clubImg = Image(Point(580, 230), "pics/clubs/" + player.getClub(num) + ".png")
            clubImg.draw(pwin)
        else:
            clubImg.undraw()
            clubImg = Image(Point(580, 230), "pics/clubs/blank.png")
            clubImg.draw(pwin)

        #Name on card
        playerNameC.undraw()
        playerNameC = Text(Point(650, 267), player.getName(num))
        playerNameC.setSize(16)
        playerNameC.setFace("helvetica")
        playerNameC.setFill(colorList[colorInd])
        playerNameC.setStyle("bold")
        playerNameC.draw(pwin)
        

        #Printed Details
        if pID:
            playerID.setText(int(player.getID(num))+1)
        playerName.setText(player.getName(num))
        playerAge.setText(player.getAge(num))
        playerNation.setText(player.getNation(num))
        playerClub.setText(player.getClub(num))
        playerPotential.setText(player.getPotential(num))
        playerValue.setText(player.getValue(num))
        playerWage.setText(player.getWage(num))
        playerFoot.setText(player.getFoot(num))
        playerWFoot.setText(str(player.getWFoot(num))+"/5")
        playerSkill.setText(str(player.getSkill(num))+"/5")
        playerIntRating.setText(str(player.getIntRating(num))+"/5")
        playerNumber.setText(player.getNumber(num))
        playerHeight.setText(player.getHeight(num))
        playerWeight.setText(player.getWeight(num))
        playerRelease.setText(player.getRelease(num))
        
        try:
            pt = pwin.getMouse()
        except GraphicsError:
            return

    pwin.close()    


def databaseCheck():
    
    #Creating file if it doesn't exist
    try:
        wb = load_workbook("userInfo.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        wb.create_sheet("User Info")
        ws = wb["User Info"]

        ws["A1"] = "Username"
        ws["B1"] = "PIN"
            
    wb.save("userInfo.xlsx")
    wb.close() 
    
def check(name, pin):

    #Boolean Checker initialization
    nameExists = False
    passwordMatch = False
    
    #Checking if login information is correct

    #Making Sheet active
    ws = wb["User Info"]

    #Checking if name Exists in the database
    for nCell in ws["A"]:
        if nCell.value.lower() == name.lower():
            nameExists = True

    #Defining the next available Row
    emptyRow = ws.max_row + 1

    #Checking if name exists
    if (nameExists):
        for nCell in ws["A"]:
            for pCell in ws["B"]:
                if nCell.row == pCell.row:
                    if ((name.lower() + pin) == (nCell.value.lower() + pCell.value)):
                        passwordMatch = True
    
    else:
        ws.cell(row = emptyRow, column = 1).value = name
        ws.cell(row = emptyRow, column = 2).value = pin

        #Create sheet for player
        wb.create_sheet(name)

        #Making sheet active
        ws = wb[name]

        #Adding information to sheet
        ws["A1"] = "Inventory"
        ws["A2"] = "Balance"

        #Starting players get balance of 200 million Euros
        ws["B2"] = 200000000
        ws["A3"] = "ID of Player"
        ws["B3"] = "Name of Player"
        ws["C3"] = "Paid Value"

        #Saving and closing sheet
        saveBook()
        
        return True

    #If both name exists and the password matches, log the user in with boolean True
    if nameExists and passwordMatch:
        return True
    else:
        return False

#Userdata save and close function
def saveBook():
    wb.save("userInfo.xlsx")

Intro()
